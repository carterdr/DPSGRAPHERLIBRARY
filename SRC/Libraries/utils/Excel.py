import openpyxl
import matplotlib.pyplot as plt
import os
"""
Excel Utilities for Damage Simulation Output

This module provides functions to read from, write to, and display data in a central Excel file (`Dps.xlsx`).
It supports naming columns, filling time rows, exporting damage data, and visualizing results via matplotlib.

Functions:
    - get_excel_Path(): Returns the full path to the `Dps.xlsx` file.
    - rename_column(col, name): Renames the header of a specific column.
    - get_open_column(sheet): Finds the next available column in the sheet.
    - create_time(): Initializes the first column of the sheet with time values from 0.0 to 100.0.
    - print_to_sheet(damageResult, custom_name): Writes a DamageResult to a new column in the Excel sheet.
    - clear_excel(): Clears all cell values from the Excel sheet.
    - display_data(): Plots all damage columns using matplotlib and saves as a PNG.
    - get_final_damage_row(sheet, final_column): Estimates the last meaningful row of damage data for display.
"""
def rename_column(col, name):
    """
    Renames the header cell of a specific column in the Excel sheet.

    Args:
        col (int): The column number (1-indexed).
        name (str): The new name for the column.
    """
    file_path = get_excel_Path()
    wb = openpyxl.load_workbook(file_path)
    sheet = wb["Sheet1"]
    sheet.cell(1,col).value = name
    wb.save(file_path)
def get_open_column(sheet):
    """
    Finds the next available column in the Excel sheet by scanning the first row.

    Args:
        sheet (Worksheet): The active Excel worksheet.

    Returns:
        int: The next open column number.
    """
    column = 2         
    while(not sheet.cell(1,column).value==None):
        column+=1  
    return column    
def get_excel_Path():
    """
    Constructs the full path to `Dps.xlsx` relative to this script's location.

    Returns:
        str: Full file path to `Dps.xlsx`.
    """
    caller_directory = os.path.dirname(os.path.realpath(__file__))
    # Move up one directory to reach the parent directory (SRC)
    project_directory = os.path.dirname(caller_directory)
    # Construct the full path to Dps.xlsx
    return os.path.join(project_directory, "..\\Dps.xlsx")

def create_time():
    """
    Writes 0.0 to 100.0 (step 0.1) in the first column under the "Seconds" header.
    """
    file_path = get_excel_Path()
    wb = openpyxl.load_workbook(file_path)
    sh1 = wb["Sheet1"]
    sh1.cell(row=1, column=1, value="Seconds")
    for row in range(1001):
        sh1.cell(row=row+2, column=1).value = "{:.1f}".format(row/10)           
    wb.save(file_path)
    
def print_to_sheet(damageResult, custom_name = None):
    """
    Writes a DamageResult to the next open column in the Excel sheet.

    Args:
        damageResult (DamageResult): The result object containing time-indexed damage values.
        custom_name (str, optional): If provided, overrides the column header name.

    Returns:
        int: The column number where data was written.
    """
    name = damageResult.name
    damage = damageResult.dot
    if custom_name != None:
        name = custom_name
    file_path = get_excel_Path()
    wb = openpyxl.load_workbook(file_path)
    sh1 = wb["Sheet1"]
    column = get_open_column(sh1)
    sh1.cell(1,column,name)    
    for i in range(0,1001):
        value = damage[i]
        sh1.cell(i+2, column).value = value
    wb.save(file_path)
    return column
    


def clear_excel():
    """
    Clears all contents of 'Sheet1' in `Dps.xlsx`.
    """
    print(get_excel_Path())
    file_path = get_excel_Path()
    book = openpyxl.load_workbook(file_path) #get the file name
    sheet = book.get_sheet_by_name('Sheet1') #get the sheet name
    for a in sheet.iter_cols(): #you can set the range here 
        for cell in a:
            cell.value = None #set a value or null here
    book.save(file_path)
def display_data():
    """
    Reads all damage columns from Excel, generates a step plot, and saves it as 'damage_over_time_graph.png'.
    """
    file_path = get_excel_Path()
    # Load the workbook and select the active sheet
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    # Read the 'seconds' column (assuming it is the first column)
    open_column = get_open_column(sheet)
    max_row = get_final_damage_row(sheet, open_column)
    seconds = [float(sheet.cell(row=i, column=1).value) for i in range(2, max_row + 1)]

    # Create a figure and an axis object
    fig, ax = plt.subplots(figsize=(15, 8))
    
    for column in range(2, open_column):
        damage = [float(sheet.cell(row=i, column=column).value) for i in range(2, max_row + 1)]
        name = sheet.cell(1, column).value
        ax.step(seconds, damage, where='post', label=name, linewidth=2)

    # Setting the labels and title
    ax.set_xlabel('Time (Seconds)', fontsize=14)
    ax.set_ylabel('Damage', fontsize=14)
    ax.set_title('Damage Over Time', fontsize=16)

    # Setting the tick parameters for the x and y axes
    ax.tick_params(axis='x', which='both')
    ax.tick_params(axis='y', which='both')

    # Setting the background color
    ax.set_facecolor('white')  # Change the background to white if preferred

    # Creating a grid
    ax.grid(color='gray', linestyle='--', linewidth=0.5)  # Add a gray dashed grid

    # Adding a legend
    ax.legend()

    # Saving the figure
    plt.savefig('damage_over_time_graph.png')  # Save the figure as a .png file

    # Show the plot
    plt.show()
def get_final_damage_row(sheet, final_column):
    """
    Estimates the final relevant row index by detecting the last increase in damage per column.

    Args:
        sheet (Worksheet): Excel worksheet object.
        final_column (int): Column number up to which damage data exists.

    Returns:
        int: Estimated row index to use for plotting/reading data.
    """
    maxRow = 2
    for column in range(2, final_column):
        maxDamageValue = 0
        lastDamageIncreaseRow = 0
        for row in range(2, 1003):
            currentValue = sheet.cell(row, column).value
            if currentValue > maxDamageValue:
                lastDamageIncreaseRow = row
                maxDamageValue = currentValue
        if maxRow < lastDamageIncreaseRow:
            maxRow = lastDamageIncreaseRow
    maxRow = int(format(maxRow * 1.25, ".0f"))
    if maxRow > 1002:
        maxRow = 1002
    return maxRow                
def prepare_sheet(clear_sheet=True):
    """
    Prepares the Excel sheet for a new simulation run.

    Optionally clears all existing data, then writes time values (0.0 to 100.0)
    into the first column under the "Seconds" header.

    Args:
        clear_excel (bool): If True (default), clears the sheet before adding time column.
    """
    if clear_sheet:
        clear_excel()
    create_time()
    