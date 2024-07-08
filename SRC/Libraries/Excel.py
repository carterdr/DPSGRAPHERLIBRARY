import openpyxl
import matplotlib.pyplot as plt
import os
from Libraries import Weapon
def renameColumn(col, name):
    file_path = Excel.getExcelPath()
    wb = openpyxl.load_workbook(file_path)
    sheet = wb["Sheet1"]
    sheet.cell(1,col).value = name
    wb.save(file_path)

class Excel():
    def __init__ (self, name, placeInColumn = None):
        self.name = name
        self.file_path = Excel.getExcelPath()
        self.wb = openpyxl.load_workbook(self.file_path)
        self.sh1 = self.wb["Sheet1"]
        self.column = Excel.getOpenColumn(self.sh1)
        if placeInColumn != None:
            self.column = placeInColumn
            self.name = self.sh1.cell(1, self.column).value
            if name != "":
                self.name += " + " + name 
        self.sh1.cell(1,self.column,self.name)     

    def getOpenColumn(sheet):
        column = 2         
        while(not sheet.cell(1,column).value==None):
            if(not sheet.cell(1,column).value==None):
                column+=1  
        return column    
    def getExcelPath():
        caller_directory = os.path.dirname(os.path.realpath(__file__))
        # Move up one directory to reach the parent directory (SRC)
        project_directory = os.path.dirname(caller_directory)
        # Construct the full path to Dps.xlsx
        return os.path.join(project_directory, "Dps.xlsx")
    def createTime():
        file_path = Excel.getExcelPath()
        wb = openpyxl.load_workbook(file_path)
        sh1 = wb["Sheet1"]
        sh1.cell(row=1, column=1, value="Seconds")
        for row in range(1001):
            sh1.cell(row=row+2, column=1).value = "{:.1f}".format(row/10)           
        wb.save(file_path)
    def closeExcel(self, damagetimes):
        damagetimes = self._removeDupeValues(damagetimes)
        cumulative_damage = 0 


        if damagetimes:
            first_time = damagetimes[0][0] + 1
        else:
            first_time = 1002  # If there are no damage times, set to the last row

        # Initialize column with zeros up to the first damage time
        for row in range(2, first_time + 1):
            if self.sh1.cell(row, self.column).value is None:
                self.sh1.cell(row, self.column).value = 0
                
        for i in range(0,len(damagetimes)):
            current_time, value = damagetimes[i]
            cumulative_damage += value
            if i == len(damagetimes)-1:
                next_time = 1002
            else:
                next_time = damagetimes[i+1][0]
            for row in range(current_time + 1, next_time + 1):  # Update from the previous time + 1 to the current time + 1
                if self.sh1.cell(row, self.column).value is None:
                    self.sh1.cell(row, self.column).value = 0
                self.sh1.cell(row, self.column).value += cumulative_damage
        self.wb.save(self.file_path)
        return self.column

    def _removeDupeValues(self, damagetimes):
        newTimes = {}
        #remove dupes
        for time, damage in damagetimes:
            damage = int(format(damage * Weapon.story_mission_to_raid_scalar, ".0f"))
            if time not in newTimes:
                newTimes[time] = damage 
            if newTimes[time] < damage:
                newTimes[time] = damage
        new_damage_times = []
        for time in sorted(newTimes.keys()):
            new_damage_times.append((time, newTimes[time]))
        damagetimes = new_damage_times
        #change format
        newTimes = {}
        cumulative_damage = 0    
        for time, damage in damagetimes:
            if time < 1002:
                newTimes[time] = damage - cumulative_damage
                cumulative_damage = damage
        newDamageTimes = []
        for time in sorted(newTimes.keys()):
            newDamageTimes.append((time, newTimes[time]))
        return newDamageTimes

    def clearExcel():
        print(Excel.getExcelPath())
        file_path = Excel.getExcelPath()
        book = openpyxl.load_workbook(file_path) #get the file name
        sheet = book.get_sheet_by_name('Sheet1') #get the sheet name
        for a in sheet.iter_cols(): #you can set the range here 
            for cell in a:
                cell.value = None #set a value or null here
        book.save(file_path)
    def displayData():
        file_path = Excel.getExcelPath()
        # Load the workbook and select the active sheet
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        # Read the 'seconds' column (assuming it is the first column)
        open_column = Excel.getOpenColumn(sheet)
        max_row = Excel.getFinalDamageRow(sheet, open_column)
        seconds = [float(sheet.cell(row=i, column=1).value) for i in range(2, max_row + 1)]

        # Create a figure and an axis object
        fig, ax = plt.subplots(figsize=(15, 8))
        
        for column in range(2, open_column):
            damage = [float(sheet.cell(row=i, column=column).value) for i in range(2, max_row + 1)]
            name = sheet.cell(1, column).value
            ax.step(seconds, damage, where='post', label=name, linewidth=2)

        # Setting the labels and title
        ax.set_xlabel('Time (Seconds)', fontsize=14)
        ax.set_ylabel('Damage', fontsize=14)
        ax.set_title('Damage Over Time', fontsize=16)

        # Setting the tick parameters for the x and y axes
        ax.tick_params(axis='x', which='both')
        ax.tick_params(axis='y', which='both')

        # Setting the background color
        ax.set_facecolor('white')  # Change the background to white if preferred

        # Creating a grid
        ax.grid(color='gray', linestyle='--', linewidth=0.5)  # Add a gray dashed grid

        # Adding a legend
        ax.legend()

        # Saving the figure
        plt.savefig('damage_over_time_graph.png')  # Save the figure as a .png file

        # Show the plot
        plt.show()
    def getFinalDamageRow(sheet, final_column):
        maxRow = 2
        for column in range(2, final_column):
            maxDamageValue = 0
            lastDamageIncreaseRow = 0
            for row in range(2, 1003):
                currentValue = sheet.cell(row, column).value
                if currentValue > maxDamageValue:
                    lastDamageIncreaseRow = row
                    maxDamageValue = currentValue
            if maxRow < lastDamageIncreaseRow:
                maxRow = lastDamageIncreaseRow
        maxRow = int(format(maxRow * 1.25, ".0f"))
        if maxRow > 1002:
            maxRow = 1002
        return maxRow                 
    
        