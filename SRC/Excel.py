import openpyxl
class Excel():
    def __init__ (self, name):
        self.name = name
        self.wb = openpyxl.load_workbook("F:\Coding\Dps.xlsx")
        self.sh1 = self.wb["Sheet1"]
        self.column=1
        while(not self.sh1.cell(1,self.column).value==None):
            if(not self.sh1.cell(1,self.column).value==None):
                self.column+=1    
        self.sh1.cell(1,self.column,name)                    
    def createTime(self):
        if(self.sh1.cell(1,1).value==None):
                    self.sh1.cell(1,1,"Seconds")
                    for i in range(1001):
                        self.sh1.cell(i+2,1,i/10)        
    def closeExcel (self):
        for i in range(1001):
            if(self.sh1.cell(i+2,self.column).value==None):
                if(type(self.sh1.cell(i+1,self.column).value)== type("string")):
                    self.sh1.cell(i+2,self.column,0)
                else:
                    self.sh1.cell(i+2,self.column,self.sh1.cell(i+1,self.column).value)
        self.wb.save("F:\Coding\Dps.xlsx")        
    def clearExcel():
        book = openpyxl.load_workbook("F:\Coding\Dps.xlsx") #get the file name
        sheet = book.get_sheet_by_name('Sheet1') #get the sheet name
        for a in sheet['A1':'Z1002']: #you can set the range here 
            for cell in a:
                cell.value = None #set a value or null here
        book.save("F:\Coding\Dps.xlsx")